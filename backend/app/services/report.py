import csv
from datetime import date, datetime
import io
from typing import List, Optional
import uuid
from sqlalchemy import and_, select, func, case
from sqlalchemy.ext.asyncio import AsyncSession

# Models
from app.models.attendance import Attendance
from app.models.student import Student
from app.models.class_session import ClassSession
from app.models.subject import Subject

async def query_attendance_report_data(
    db: AsyncSession,
    class_id: Optional[int] = None,
    subject_id: Optional[int] = None,
    student_id: Optional[uuid.UUID] = None,
    department: Optional[str] = None,
    year: Optional[int] = None,
    section: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> List[tuple]:
    """
    Shared helper method to query attendance database logs joined with 
    Student, ClassSession, and Subject records matching parameters.
    """
    query = select(Attendance, Student, ClassSession, Subject).join(
        Student, Attendance.student_id == Student.id
    ).join(
        ClassSession, Attendance.class_id == ClassSession.id
    ).join(
        Subject, ClassSession.subject_id == Subject.id
    )
    
    conditions = []
    if class_id:
        conditions.append(Attendance.class_id == class_id)
    if subject_id:
        conditions.append(ClassSession.subject_id == subject_id)
    if student_id:
        conditions.append(Attendance.student_id == student_id)
    if department:
        conditions.append(Student.department == department)
    if year:
        conditions.append(Student.year == year)
    if section:
        conditions.append(Student.section == section)
    if start_date:
        start_dt = datetime.combine(start_date, datetime.min.time())
        conditions.append(Attendance.marked_at >= start_dt)
    if end_date:
        end_dt = datetime.combine(end_date, datetime.max.time())
        conditions.append(Attendance.marked_at <= end_dt)
        
    if conditions:
        query = query.where(and_(*conditions))
        
    # Sort chronologically descending
    query = query.order_by(Attendance.marked_at.desc())
    
    result = await db.execute(query)
    return list(result.all())


def generate_csv_report(data: List[tuple]) -> str:
    """Generates standard CSV comma-delimited output content."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Marked At", "Roll Number", "Name", "Department", "Year", "Section",
        "Subject Code", "Subject Name", "Classroom", "Status", "Method", "Confidence", "Flagged"
    ])
    
    for att, student, cls, subj in data:
        writer.writerow([
            att.marked_at.strftime("%Y-%m-%d %H:%M:%S") if att.marked_at else "",
            student.roll_number,
            student.name,
            student.department,
            student.year,
            student.section,
            subj.code,
            subj.name,
            cls.classroom,
            att.status.upper(),
            att.method,
            f"{att.confidence_score:.2f}" if att.confidence_score is not None else "N/A",
            "YES" if att.is_flagged else "NO"
        ])
        
    return output.getvalue()


def generate_excel_report(data: List[tuple]) -> bytes:
    """Generates a professional styled Excel spreadsheet using OpenPyXL."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"
    
    # Theme palettes
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    accent_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border_side = Side(border_style="thin", color="E0E0E0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Title Block
    ws.merge_cells("A1:M1")
    ws["A1"] = "SmartAttend AI — Attendance Verification Sheet"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws["A2"] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)
    ws.row_dimensions[2].height = 20
    
    # Headers
    headers = [
        "Marked At", "Roll Number", "Name", "Department", "Year", "Section",
        "Subject Code", "Subject Name", "Classroom", "Status", "Method", "Confidence", "Flagged"
    ]
    
    ws.row_dimensions[4].height = 25
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # Write entries
    current_row = 5
    for att, student, cls, subj in data:
        ws.row_dimensions[current_row].height = 20
        
        row_values = [
            att.marked_at.strftime("%Y-%m-%d %H:%M") if att.marked_at else "",
            student.roll_number,
            student.name,
            student.department,
            student.year,
            student.section,
            subj.code,
            subj.name,
            cls.classroom,
            att.status.upper(),
            att.method.capitalize(),
            att.confidence_score if att.confidence_score is not None else "N/A",
            "Yes" if att.is_flagged else "No"
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1, 2, 4, 5, 6, 7, 9, 10, 11, 12, 13]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            # Zebra striping
            if current_row % 2 == 0:
                cell.fill = accent_fill
                
            # Formatting Status / Flags
            if col_idx == 10:  # Status
                if val == "PRESENT":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="2E7D32")
                elif val == "ABSENT":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="C62828")
                else:
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="EF6C00")
            elif col_idx == 13 and val == "Yes":  # Flagged
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="C62828")
                cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                
        current_row += 1
        
    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_pdf_report(data: List[tuple]) -> bytes:
    """Generates styled landscape PDF document via ReportLab flowables."""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1B365D'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        spaceAfter=15
    )
    card_title_style = ParagraphStyle(
        'CardTitle',
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    card_value_style = ParagraphStyle(
        'CardValue',
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.white,
        alignment=1
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        alignment=0
    )
    table_cell_center = ParagraphStyle(
        'TableCellCenter',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        alignment=1
    )
    
    elements = []
    
    elements.append(Paragraph("SmartAttend AI — Attendance Summary Report", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
    
    # Summarize stats
    total_records = len(data)
    present_records = sum(1 for att, _, _, _ in data if att.status in ["present", "late"])
    flagged_records = sum(1 for att, _, _, _ in data if att.is_flagged)
    
    avg_percentage = (present_records / total_records) * 100.0 if total_records > 0 else 100.0
    flagged_rate = (flagged_records / total_records) * 100.0 if total_records > 0 else 0.0
    
    # Stat Boxes Block
    summary_data = [
        [
            Paragraph("Total Logs Checked", card_title_style),
            Paragraph("Average Present %", card_title_style),
            Paragraph("Flagged Anomalies Count", card_title_style)
        ],
        [
            Paragraph(str(total_records), card_value_style),
            Paragraph(f"{avg_percentage:.1f}%", card_value_style),
            Paragraph(f"{flagged_records} ({flagged_rate:.1f}%)", card_value_style)
        ]
    ]
    summary_table = Table(summary_data, colWidths=[240, 240, 240])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#1B365D')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 1.5, colors.white),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Table Grid
    table_headers = [
        Paragraph("<b>Marked At</b>", table_cell_center),
        Paragraph("<b>Roll Number</b>", table_cell_center),
        Paragraph("<b>Name</b>", table_cell_style),
        Paragraph("<b>Dept.</b>", table_cell_center),
        Paragraph("<b>Yr/Sec</b>", table_cell_center),
        Paragraph("<b>Subject Code/Name</b>", table_cell_style),
        Paragraph("<b>Classroom</b>", table_cell_center),
        Paragraph("<b>Status</b>", table_cell_center),
        Paragraph("<b>Method</b>", table_cell_center),
        Paragraph("<b>Confidence</b>", table_cell_center),
        Paragraph("<b>Flagged</b>", table_cell_center)
    ]
    
    table_rows = [table_headers]
    for att, student, cls, subj in data:
        status_color = "#2E7D32" if att.status in ["present", "late"] else "#C62828"
        status_html = f"<font color='{status_color}'><b>{att.status.upper()}</b></font>"
        flagged_html = "<font color='#C62828'><b>Yes</b></font>" if att.is_flagged else "No"
        
        table_rows.append([
            Paragraph(att.marked_at.strftime("%Y-%m-%d %H:%M") if att.marked_at else "", table_cell_center),
            Paragraph(student.roll_number, table_cell_center),
            Paragraph(student.name, table_cell_style),
            Paragraph(student.department, table_cell_center),
            Paragraph(f"{student.year}/{student.section}", table_cell_center),
            Paragraph(f"{subj.code} - {subj.name}", table_cell_style),
            Paragraph(cls.classroom, table_cell_center),
            Paragraph(status_html, table_cell_center),
            Paragraph(att.method.capitalize(), table_cell_center),
            Paragraph(f"{att.confidence_score:.2f}" if att.confidence_score is not None else "N/A", table_cell_center),
            Paragraph(flagged_html, table_cell_center)
        ])
        
    col_widths = [90, 75, 95, 35, 40, 130, 60, 50, 55, 60, 50]
    logs_table = Table(table_rows, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EAEAEA')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D3D3D3')),
    ])
    
    for i in range(1, len(table_rows)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F7FA'))
            
    logs_table.setStyle(t_style)
    elements.append(logs_table)
    
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawString(36, 18, "CONFIDENTIAL — SmartAttend AI System")
        page_num = canvas.getPageNumber()
        canvas.drawRightString(landscape(letter)[0] - 36, 18, f"Page {page_num}")
        canvas.restoreState()
        
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    return buffer.getvalue()


async def query_overall_attendance(
    db: AsyncSession,
    department: Optional[str] = None,
    year: Optional[int] = None,
    section: Optional[str] = None,
    subject_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> List[dict]:
    """
    Groups and aggregates student logs to compute attendance percents and risk status.
    Runs as ONE unified grouped SQL query.
    """
    att_conditions = []
    use_class_session = (subject_id is not None)
    
    att_query = select(
        Attendance.student_id.label("student_id"),
        func.count(Attendance.id).label("total"),
        func.sum(
            case(
                (Attendance.status.in_(["present", "late"]), 1),
                else_=0
            )
        ).label("attended")
    )
    
    if use_class_session:
        att_query = att_query.join(ClassSession, Attendance.class_id == ClassSession.id)
        att_conditions.append(ClassSession.subject_id == subject_id)
        
    if start_date:
        start_dt = datetime.combine(start_date, datetime.min.time())
        att_conditions.append(Attendance.marked_at >= start_dt)
    if end_date:
        end_dt = datetime.combine(end_date, datetime.max.time())
        att_conditions.append(Attendance.marked_at <= end_dt)
        
    if att_conditions:
        att_query = att_query.where(and_(*att_conditions))
        
    att_subquery = att_query.group_by(Attendance.student_id).subquery()
    
    student_conditions = [Student.is_active == True]
    if department:
        student_conditions.append(Student.department == department)
    if year:
        student_conditions.append(Student.year == year)
    if section:
        student_conditions.append(Student.section == section)
        
    main_query = select(
        Student.id.label("student_uid"),
        Student.roll_number,
        Student.name.label("full_name"),
        Student.department,
        Student.year,
        Student.section,
        func.coalesce(att_subquery.c.total, 0).label("total_sessions"),
        func.coalesce(att_subquery.c.attended, 0).label("attended_sessions")
    ).outerjoin(
        att_subquery, Student.id == att_subquery.c.student_id
    ).where(
        and_(*student_conditions)
    ).order_by(
        Student.roll_number.asc()
    )
    
    result = await db.execute(main_query)
    rows = result.all()
    
    report_data = []
    for r in rows:
        total = int(r.total_sessions)
        attended = int(r.attended_sessions)
        pct = float((attended / total * 100.0) if total > 0 else 0.0)
        
        if total == 0:
            risk = "NO_DATA"
        elif pct >= 85.0:
            risk = "SAFE"
        elif pct >= 75.0:
            risk = "WARNING"
        else:
            risk = "CRITICAL"
            
        report_data.append({
            "student_uid": r.student_uid,
            "roll_number": r.roll_number,
            "full_name": r.full_name,
            "department": r.department,
            "year": r.year,
            "section": r.section,
            "total_sessions": total,
            "attended_sessions": attended,
            "attendance_percent": pct,
            "risk_status": risk
        })
        
    return report_data


def generate_overall_csv_report(data: List[dict]) -> str:
    """Generates standard CSV output content for overall attendance."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        "Roll Number", "Full Name", "Department", "Year", "Section",
        "Total Sessions", "Attended Sessions", "Attendance %", "Risk Status"
    ])
    
    for row in data:
        writer.writerow([
            row["roll_number"],
            row["full_name"],
            row["department"],
            row["year"],
            row["section"],
            row["total_sessions"],
            row["attended_sessions"],
            f"{row['attendance_percent']:.1f}%",
            row["risk_status"]
        ])
        
    return output.getvalue()


def generate_overall_excel_report(data: List[dict]) -> bytes:
    """Generates a professional styled Excel spreadsheet for overall attendance using OpenPyXL."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Overall Attendance"
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    accent_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border_side = Side(border_style="thin", color="E0E0E0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    ws.merge_cells("A1:I1")
    ws["A1"] = "SmartAttend AI — Overall Attendance Summary Sheet"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws["A2"] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)
    ws.row_dimensions[2].height = 20
    
    headers = [
        "Roll Number", "Full Name", "Department", "Year", "Section",
        "Total Sessions", "Attended Sessions", "Attendance %", "Risk Status"
    ]
    
    ws.row_dimensions[4].height = 25
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    current_row = 5
    for row in data:
        ws.row_dimensions[current_row].height = 20
        
        row_values = [
            row["roll_number"],
            row["full_name"],
            row["department"],
            row["year"],
            row["section"],
            row["total_sessions"],
            row["attended_sessions"],
            f"{row['attendance_percent']:.1f}%",
            row["risk_status"]
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border
            
            if col_idx in [1, 4, 5, 6, 7, 8, 9]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            if current_row % 2 == 0:
                cell.fill = accent_fill
                
            if col_idx == 9:
                if val == "SAFE":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="2E7D32")
                elif val == "WARNING":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="EF6C00")
                elif val == "CRITICAL":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="C62828")
                else:
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="616161")
                
        current_row += 1
        
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_overall_pdf_report(data: List[dict]) -> bytes:
    """Generates styled landscape PDF document for overall attendance via ReportLab."""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1B365D'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        spaceAfter=15
    )
    card_title_style = ParagraphStyle(
        'CardTitle',
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    card_value_style = ParagraphStyle(
        'CardValue',
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.white,
        alignment=1
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        alignment=0
    )
    table_cell_center = ParagraphStyle(
        'TableCellCenter',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        alignment=1
    )
    
    elements = []
    
    elements.append(Paragraph("SmartAttend AI — Overall Attendance Summary Report", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
    
    total_students = len(data)
    critical_count = sum(1 for row in data if row["risk_status"] == "CRITICAL")
    warning_count = sum(1 for row in data if row["risk_status"] == "WARNING")
    safe_count = sum(1 for row in data if row["risk_status"] == "SAFE")
    
    summary_data = [
        [
            Paragraph("Total Students", card_title_style),
            Paragraph("Safe (>=85%)", card_title_style),
            Paragraph("Warning (75-85%)", card_title_style),
            Paragraph("Critical (<75%)", card_title_style)
        ],
        [
            Paragraph(str(total_students), card_value_style),
            Paragraph(str(safe_count), card_value_style),
            Paragraph(str(warning_count), card_value_style),
            Paragraph(str(critical_count), card_value_style)
        ]
    ]
    summary_table = Table(summary_data, colWidths=[180, 180, 180, 180])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#1B365D')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 1.5, colors.white),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    table_headers = [
        Paragraph("<b>Roll Number</b>", table_cell_center),
        Paragraph("<b>Full Name</b>", table_cell_style),
        Paragraph("<b>Dept.</b>", table_cell_center),
        Paragraph("<b>Yr</b>", table_cell_center),
        Paragraph("<b>Sec</b>", table_cell_center),
        Paragraph("<b>Total Sessions</b>", table_cell_center),
        Paragraph("<b>Attended Sessions</b>", table_cell_center),
        Paragraph("<b>Attendance %</b>", table_cell_center),
        Paragraph("<b>Risk Status</b>", table_cell_center)
    ]
    
    table_rows = [table_headers]
    for row in data:
        status = row["risk_status"]
        if status == "SAFE":
            status_color = "#2E7D32"
        elif status == "WARNING":
            status_color = "#EF6C00"
        elif status == "CRITICAL":
            status_color = "#C62828"
        else:
            status_color = "#616161"
            
        status_html = f"<font color='{status_color}'><b>{status}</b></font>"
        
        table_rows.append([
            Paragraph(row["roll_number"], table_cell_center),
            Paragraph(row["full_name"], table_cell_style),
            Paragraph(row["department"], table_cell_center),
            Paragraph(str(row["year"]), table_cell_center),
            Paragraph(row["section"], table_cell_center),
            Paragraph(str(row["total_sessions"]), table_cell_center),
            Paragraph(str(row["attended_sessions"]), table_cell_center),
            Paragraph(f"{row['attendance_percent']:.1f}%", table_cell_center),
            Paragraph(status_html, table_cell_center)
        ])
        
    col_widths = [80, 150, 80, 30, 30, 90, 90, 80, 110]
    logs_table = Table(table_rows, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EAEAEA')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D3D3D3')),
    ])
    
    for i in range(1, len(table_rows)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F7FA'))
            
    logs_table.setStyle(t_style)
    elements.append(logs_table)
    
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawString(36, 18, "CONFIDENTIAL — SmartAttend AI Overall Report")
        page_num = canvas.getPageNumber()
        canvas.drawRightString(landscape(letter)[0] - 36, 18, f"Page {page_num}")
        canvas.restoreState()
        
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    return buffer.getvalue()

